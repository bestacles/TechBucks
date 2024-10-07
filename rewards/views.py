import logging  # Import the logging module
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from openpyxl import load_workbook  # For handling XLSX files
from .models import Reward, Redemption
from students.models import Student
from .forms import XLSXUploadForm
from django.db import transaction

# Create logger
logger = logging.getLogger(__name__)

# View to handle reward creation via form
def create_reward(request):
    if request.method == 'POST':
        try:
            reward = Reward.objects.create(
                name=request.POST['name'],
                description=request.POST['description'],
                cost=request.POST['cost']
            )
            reward.generate_qr_code()  # Call to generate and save QR code

            messages.success(request, "Reward created successfully!")
            return redirect('rewards:reward_list')
        except Exception as e:
            logger.error(f"Error creating reward: {str(e)}")
            messages.error(request, f"Error creating reward: {str(e)}")
            return redirect('rewards:error_page')
    return render(request, 'rewards/reward_form.html')


# View to list rewards with optional sorting
def reward_list(request):
    try:
        sort_by = request.GET.get('sort', 'name')  # Default sort by name
        rewards = Reward.objects.all().order_by(sort_by)
        return render(request, 'rewards/reward_list.html', {'rewards': rewards})
    except Exception as e:
        logger.error(f"Error listing rewards: {str(e)}")
        messages.error(request, f"Error listing rewards: {str(e)}")
        return redirect('rewards:error_page')


# View to handle XLSX file upload and adding rewards
def add_reward_view(request):
    if request.method == 'POST':
        form = XLSXUploadForm(request.POST, request.FILES)
        if form.is_valid():
            xlsx_file = request.FILES['xlsx_file']

            try:
                logger.info(f"Attempting to load XLSX file: {xlsx_file.name}")
                workbook = load_workbook(xlsx_file)
                sheet = workbook.active

                logger.info(f"Workbook loaded successfully. Processing sheet...")
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                    logger.info(f"Processing row: {row}")
                    Reward.objects.create(
                        name=row[0],
                        description=row[1],
                        cost=int(row[2]),
                        category=row[3]
                    )
                    logger.info(f"Reward {row[0]} added successfully.")

                messages.success(request, "Rewards added successfully!")
                return redirect('rewards:reward_list')

            except Exception as e:
                logger.error(f"Error processing XLSX file: {str(e)}")
                messages.error(request, f"Error processing XLSX file: {str(e)}")
                return redirect('rewards:error_page')

    else:
        form = XLSXUploadForm()

    return render(request, 'rewards/add_reward.html', {'form': form})


# View to handle reward redemption
from transactions.models import Transaction  # Import the Transaction model if you haven't already

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from transactions.models import Transaction
from .models import Reward, Redemption
from students.models import Student
import logging

logger = logging.getLogger(__name__)


@transaction.atomic
def redeem_reward(request, student_id, reward_id):
    student = get_object_or_404(Student, student_id=str(student_id))
    reward = get_object_or_404(Reward, id=reward_id)

    logger.debug(f"Redeeming reward for student {student.student_id}, reward {reward.id}")

    if request.method == 'POST':
        logger.debug(f"Student TechBucks before transaction: {student.techbucks}, Reward cost: {reward.cost}")

        # Check if student has enough TechBucks
        if student.techbucks >= reward.cost:
            # Create a transaction to deduct TechBucks
            Transaction.objects.create(
                student=student,
                amount=-reward.cost,  # Deduct reward cost
                description=f"Redeemed {reward.name}",
                performed_by=request.user  # Track the user who performed the transaction
            )

            # Create a Redemption entry
            Redemption.objects.create(student=student, reward=reward)

            # Success message
            messages.success(request, f"Successfully redeemed {reward.name} for {reward.cost} TechBucks!")
            logger.debug(f"Student TechBucks after transaction: {student.techbucks - reward.cost}")
            return redirect('students:student_profile', student_id=student.student_id)
        else:
            messages.error(request, "Not enough TechBucks to redeem this reward.")
            logger.debug(f"Redemption failed. Not enough TechBucks: {student.techbucks}")
            return redirect('students:student_profile', student_id=student.student_id)

    return redirect('rewards:error_page')


# Success page for reward actions
def success(request):
    return render(request, 'rewards/success.html')


# Simple error page view
def error_page(request):
    return render(request, 'rewards/error_page.html')  # Simple error page
